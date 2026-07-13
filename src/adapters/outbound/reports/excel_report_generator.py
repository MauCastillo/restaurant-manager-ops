import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.ports.outbound.report_generator_port import ReportGeneratorPort


class OpenpyxlReportGenerator(ReportGeneratorPort):
    """Openpyxl implementation of ReportGeneratorPort matching reference Excel format."""

    def generate_excel_report(self, rows: List[Dict[str, Any]], headers: List[str]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja 1"

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # Row 1 is spacer (matching reference workbook)
        # Row 2 holds headers
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rows 3..N hold client data
        keys = ["nombre", "cedula", "proceso_desempeno", "valor_a_descontar"]
        current_row = 3
        for r_data in rows:
            for col_idx, key in enumerate(keys, start=1):
                val = r_data.get(key, "")
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = thin_border

                if key == "valor_a_descontar":
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif key == "cedula":
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            current_row += 1

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
